from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file
from datetime import date, datetime
import os
import io
from werkzeug.utils import secure_filename
import validators as v
from database import (
    init_db,
    # Користувачі
    create_user, get_user_by_login,
    # Студенти
    create_student, get_all_students, get_student_by_user_id,
    get_student_by_id, update_student,
    # Викладачі
    create_teacher, get_all_teachers, get_teacher_by_user_id,
    # Компанії
    create_company, get_all_companies, get_company_by_id,
    update_company, delete_company,
    # Практики
    create_practice, get_all_practices, get_practice_by_id,
    get_practice_by_student_id, get_practices_by_teacher_id,
    update_practice_status, delete_practice,
    # Щоденник
    add_diary_entry, get_diary_entries,
    update_diary_entry, delete_diary_entry,
    # Звіти
    submit_report, get_report_by_practice_id,
    grade_report, get_all_reports, return_report_for_revision,
    # Видалення
    delete_user,
    # Нові функції (тиждень 1)
    get_students_without_practice,
    search_students,
    get_practices_filtered,
    # Управління викладачами та обліковими записами
    get_all_teachers_with_logins,
    update_user_login,
    update_user_password,
    delete_teacher,
    get_teacher_by_id,
    # Аватарки
    add_avatar_column,
    update_user_avatar,
)
from auth import (
    hash_password, check_password,
    login_user, logout_user,
    login_required, admin_required,
    teacher_required, student_required,
    get_current_user_id, get_current_user_role,
)

# ============================================================
# НАЛАШТУВАННЯ ЗАСТОСУНКУ
# ============================================================

app = Flask(__name__)

# Секретний ключ для шифрування сесій.
# У реальному проєкті заміни на довгий випадковий рядок!
app.secret_key = 'praktyka-poltava-2026-secret-key'

# Максимальний розмір файлу для завантаження — 16 МБ
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Папка для аватарок
AVATAR_FOLDER = os.path.join(os.path.dirname(__file__), 'static', 'avatars')
os.makedirs(AVATAR_FOLDER, exist_ok=True)
app.config['AVATAR_FOLDER'] = AVATAR_FOLDER
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

# Папка для файлів звітів студентів
REPORT_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads', 'reports')
os.makedirs(REPORT_FOLDER, exist_ok=True)
app.config['REPORT_FOLDER'] = REPORT_FOLDER
ALLOWED_REPORT_EXTENSIONS = {'pdf', 'doc', 'docx', 'png', 'jpg', 'jpeg', 'zip', 'rar'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def allowed_report_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_REPORT_EXTENSIONS


# ============================================================
# ГОЛОВНА СТОРІНКА — ПЕРЕНАПРАВЛЕННЯ
# ============================================================

@app.route('/')
def index():
    """Головна сторінка — перенаправляє залежно від ролі."""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))


# ============================================================
# ВХІД / ВИХІД
# ============================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    GET  — показує форму входу
    POST — перевіряє логін і пароль
    """
    # Якщо вже залогінений — одразу на дашборд
    if 'user_id' in session:
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        login_input = request.form.get('login', '').strip()
        password_input = request.form.get('password', '').strip()

        # Перевіряємо чи заповнені поля
        if not login_input or not password_input:
            flash('Заповніть усі поля.', 'warning')
            return render_template('login.html')

        # Шукаємо користувача в БД
        user = get_user_by_login(login_input)

        if user and check_password(password_input, user['password_hash']):
            # Пароль правильний — зберігаємо в сесію
            login_user(user)
            session['user_avatar'] = user['avatar'] if user['avatar'] else ''
            flash(f'Ласкаво просимо, {user["login"]}!', 'success')
            return redirect(url_for('dashboard'))
        else:
            flash('Невірний логін або пароль.', 'danger')

    return render_template('login.html')


@app.route('/logout')
def logout():
    """Виходить з акаунту та перенаправляє на вхід."""
    logout_user()
    flash('Ви вийшли з системи.', 'info')
    return redirect(url_for('login'))


# ============================================================
# ДАШБОРД — ПЕРЕНАПРАВЛЕННЯ ЗА РОЛЛЮ
# ============================================================

@app.route('/dashboard')
@login_required
def dashboard():
    """Перенаправляє кожну роль на свою головну сторінку."""
    role = get_current_user_role()
    if role == 'admin':
        return redirect(url_for('admin_dashboard'))
    elif role == 'teacher':
        return redirect(url_for('teacher_dashboard'))
    elif role == 'student':
        return redirect(url_for('student_dashboard'))
    return redirect(url_for('login'))


# ============================================================
# АДМІНІСТРАТОР
# ============================================================

@app.route('/admin')
@admin_required
def admin_dashboard():
    """Головна панель адміністратора зі статистикою."""
    students = get_all_students()
    practices = get_all_practices()
    companies = get_all_companies()
    teachers = get_all_teachers()
    reports = get_all_reports()

    # Підраховуємо статистику
    stats = {
        'total_students':  len(students),
        'total_practices': len(practices),
        'total_companies': len(companies),
        'total_teachers':  len(teachers),
        'graded':    sum(1 for p in practices if p['status'] == 'graded'),
        'submitted': sum(1 for p in practices if p['status'] == 'submitted'),
        'in_progress': sum(1 for p in practices if p['status'] == 'in_progress'),
        'assigned':  sum(1 for p in practices if p['status'] == 'assigned'),
    }

    # --- Дані для графіків ---

    # 1. Розподіл оцінок (для оцінених звітів)
    grade_buckets = {'2 (0-59)': 0, '3 (60-74)': 0, '4 (75-89)': 0, '5 (90-100)': 0}
    for r in reports:
        g = r['grade']
        if g is None:
            continue
        if g < 60:
            grade_buckets['2 (0-59)'] += 1
        elif g < 75:
            grade_buckets['3 (60-74)'] += 1
        elif g < 90:
            grade_buckets['4 (75-89)'] += 1
        else:
            grade_buckets['5 (90-100)'] += 1

    # 2. Кількість практик по групах
    groups = {}
    for p in practices:
        grp = p['student_group']
        groups[grp] = groups.get(grp, 0) + 1

    chart_data = {
        'statuses': {
            'labels': ['Призначено', 'В процесі', 'Звіт здано', 'Оцінено'],
            'values': [stats['assigned'], stats['in_progress'],
                       stats['submitted'], stats['graded']],
        },
        'grades': {
            'labels': list(grade_buckets.keys()),
            'values': list(grade_buckets.values()),
        },
        'groups': {
            'labels': list(groups.keys()),
            'values': list(groups.values()),
        },
    }

    return render_template('admin/dashboard.html', stats=stats, chart_data=chart_data)


# --- Студенти (адмін) ---

@app.route('/admin/students')
@admin_required
def admin_students():
    """Список усіх студентів з пошуком за іменем або групою."""
    query    = request.args.get('q', '').strip()
    students = search_students(query) if query else get_all_students()
    return render_template('admin/students.html', students=students, query=query)


@app.route('/admin/students/add', methods=['GET', 'POST'])
@admin_required
def admin_add_student():
    """Додавання нового студента."""
    if request.method == 'POST':
        full_name  = request.form.get('full_name', '').strip()
        group_name = request.form.get('group_name', '').strip()
        login      = request.form.get('login', '').strip()
        password   = request.form.get('password', '').strip()

        # Серверна валідація кожного поля
        error = (v.validate_full_name(full_name) or v.validate_group(group_name)
                 or v.validate_login(login) or v.validate_password(password))
        if error:
            flash(error, 'warning')
            return render_template('admin/add_student.html')

        # Спочатку створюємо користувача, потім профіль студента
        user_id = create_user(login, hash_password(password), 'student')
        if user_id is None:
            flash('Логін вже зайнятий. Оберіть інший.', 'danger')
            return render_template('admin/add_student.html')

        create_student(full_name, group_name, user_id)
        flash(f'Студента {full_name} успішно додано!', 'success')
        return redirect(url_for('admin_students'))

    return render_template('admin/add_student.html')


@app.route('/admin/students/edit/<int:student_id>', methods=['GET', 'POST'])
@admin_required
def admin_edit_student(student_id):
    """Редагування даних студента."""
    student = get_student_by_id(student_id)
    if not student:
        flash('Студента не знайдено.', 'danger')
        return redirect(url_for('admin_students'))

    if request.method == 'POST':
        full_name  = request.form.get('full_name', '').strip()
        group_name = request.form.get('group_name', '').strip()

        error = v.validate_full_name(full_name) or v.validate_group(group_name)
        if error:
            flash(error, 'warning')
            return render_template('admin/edit_student.html', student=student)

        update_student(student_id, full_name, group_name)
        flash('Дані студента оновлено!', 'success')
        return redirect(url_for('admin_students'))

    return render_template('admin/edit_student.html', student=student)


@app.route('/admin/students/delete/<int:user_id>')
@admin_required
def admin_delete_student(user_id):
    """Видалення студента (і його облікового запису)."""
    delete_user(user_id)
    flash('Студента видалено.', 'info')
    return redirect(url_for('admin_students'))


# --- Бази практики (адмін) ---

@app.route('/admin/companies')
@admin_required
def admin_companies():
    """Список баз практики."""
    companies = get_all_companies()
    return render_template('admin/companies.html', companies=companies)


@app.route('/admin/companies/add', methods=['GET', 'POST'])
@admin_required
def admin_add_company():
    """Додавання нової бази практики."""
    if request.method == 'POST':
        name           = request.form.get('name', '').strip()
        address        = request.form.get('address', '').strip()
        phone          = request.form.get('phone', '').strip()
        contact_person = request.form.get('contact_person', '').strip()

        error = (v.validate_text(name, "Назва", min_len=2, max_len=200)
                 or v.validate_text(address, "Адреса", min_len=3, max_len=200)
                 or v.validate_phone(phone))
        if error:
            flash(error, 'warning')
            return render_template('admin/add_company.html')

        create_company(name, address, phone, contact_person)
        flash(f'Базу практики "{name}" додано!', 'success')
        return redirect(url_for('admin_companies'))

    return render_template('admin/add_company.html')


@app.route('/admin/companies/edit/<int:company_id>', methods=['GET', 'POST'])
@admin_required
def admin_edit_company(company_id):
    """Редагування бази практики."""
    company = get_company_by_id(company_id)
    if not company:
        flash('Базу практики не знайдено.', 'danger')
        return redirect(url_for('admin_companies'))

    if request.method == 'POST':
        name           = request.form.get('name', '').strip()
        address        = request.form.get('address', '').strip()
        phone          = request.form.get('phone', '').strip()
        contact_person = request.form.get('contact_person', '').strip()

        error = (v.validate_text(name, "Назва", min_len=2, max_len=200)
                 or v.validate_text(address, "Адреса", min_len=3, max_len=200)
                 or v.validate_phone(phone))
        if error:
            flash(error, 'warning')
            return render_template('admin/edit_company.html', company=company)

        update_company(company_id, name, address, phone, contact_person)
        flash('Дані оновлено!', 'success')
        return redirect(url_for('admin_companies'))

    return render_template('admin/edit_company.html', company=company)


@app.route('/admin/companies/delete/<int:company_id>')
@admin_required
def admin_delete_company(company_id):
    """Видалення бази практики."""
    delete_company(company_id)
    flash('Базу практики видалено.', 'info')
    return redirect(url_for('admin_companies'))


# --- Практики (адмін) ---

@app.route('/admin/practices')
@admin_required
def admin_practices():
    """Список практик з фільтрацією за статусом і пошуком."""
    status  = request.args.get('status', '').strip() or None
    search  = request.args.get('q', '').strip()      or None
    practices = get_practices_filtered(status=status, search=search)
    teachers  = get_all_teachers()
    return render_template('admin/practices.html',
                           practices=practices,
                           teachers=teachers,
                           current_status=status or 'all',
                           current_search=search or '')


@app.route('/admin/practices/add', methods=['GET', 'POST'])
@admin_required
def admin_add_practice():
    """Призначення студента на практику."""
    if request.method == 'POST':
        student_id = request.form.get('student_id')
        company_id = request.form.get('company_id')
        teacher_id = request.form.get('teacher_id')
        start_date = request.form.get('start_date')
        end_date   = request.form.get('end_date')
        topic      = request.form.get('topic', '').strip()

        if not all([student_id, company_id, teacher_id, start_date, end_date, topic]):
            flash('Заповніть усі поля.', 'warning')
        elif v.validate_text(topic, "Тема", min_len=3, max_len=300):
            flash(v.validate_text(topic, "Тема", min_len=3, max_len=300), 'warning')
        elif v.validate_date_range(start_date, end_date):
            flash(v.validate_date_range(start_date, end_date), 'danger')
        else:
            # ✅ ЗАХИСТ ВІД ДУБЛЮВАННЯ
            existing = get_practice_by_student_id(int(student_id))
            if existing:
                flash('Цьому студенту вже призначено практику! Спочатку видаліть попередню.', 'danger')
            else:
                create_practice(student_id, company_id, teacher_id,
                                start_date, end_date, topic)
                flash('Практику призначено!', 'success')
                return redirect(url_for('admin_practices'))

    # ✅ Тільки студенти БЕЗ практики у списку вибору
    students  = get_students_without_practice()
    companies = get_all_companies()
    teachers  = get_all_teachers()
    return render_template('admin/add_practice.html',
                           students=students,
                           companies=companies,
                           teachers=teachers)


@app.route('/admin/practices/delete/<int:practice_id>')
@admin_required
def admin_delete_practice(practice_id):
    """Видалення практики."""
    delete_practice(practice_id)
    flash('Практику видалено.', 'info')
    return redirect(url_for('admin_practices'))


# --- Керівники (адмін) ---

@app.route('/admin/teachers/add', methods=['GET', 'POST'])
@admin_required
def admin_add_teacher():
    """Додавання нового керівника практики."""
    if request.method == 'POST':
        full_name = request.form.get('full_name', '').strip()
        login     = request.form.get('login', '').strip()
        password  = request.form.get('password', '').strip()

        error = (v.validate_full_name(full_name) or v.validate_login(login)
                 or v.validate_password(password))
        if error:
            flash(error, 'warning')
            return render_template('admin/add_teacher.html')

        user_id = create_user(login, hash_password(password), 'teacher')
        if user_id is None:
            flash('Логін вже зайнятий.', 'danger')
            return render_template('admin/add_teacher.html')

        create_teacher(full_name, user_id)
        flash(f'Керівника {full_name} додано!', 'success')
        return redirect(url_for('admin_dashboard'))

    return render_template('admin/add_teacher.html')



# --- Список викладачів (адмін) ---

@app.route('/admin/teachers')
@admin_required
def admin_teachers():
    """Список усіх керівників практики."""
    teachers = get_all_teachers_with_logins()
    return render_template('admin/teachers.html', teachers=teachers)


@app.route('/admin/teachers/delete/<int:user_id>')
@admin_required
def admin_delete_teacher(user_id):
    """Видалення керівника практики."""
    delete_teacher(user_id)
    flash('Керівника видалено.', 'info')
    return redirect(url_for('admin_teachers'))


# --- Редагування облікових записів (логін / пароль) ---

@app.route('/admin/credentials/<int:user_id>', methods=['GET', 'POST'])
@admin_required
def admin_edit_credentials(user_id):
    """Зміна логіну та пароля будь-якого користувача."""
    from database import get_user_by_id
    user = get_user_by_id(user_id)
    if not user:
        flash('Користувача не знайдено.', 'danger')
        return redirect(url_for('admin_dashboard'))

    if request.method == 'POST':
        new_login    = request.form.get('login', '').strip()
        new_password = request.form.get('password', '').strip()
        changed = []

        # Валідація логіну якщо змінюється
        if new_login and new_login != user['login']:
            err = v.validate_login(new_login)
            if err:
                flash(err, 'warning')
                return render_template('admin/edit_credentials.html', user=user)
            ok = update_user_login(user_id, new_login)
            if ok:
                changed.append('логін')
            else:
                flash(f'Логін «{new_login}» вже зайнятий. Оберіть інший.', 'danger')
                return render_template('admin/edit_credentials.html', user=user)

        # Валідація пароля якщо змінюється
        if new_password:
            err = v.validate_password(new_password)
            if err:
                flash(err, 'warning')
                return render_template('admin/edit_credentials.html', user=user)
            update_user_password(user_id, hash_password(new_password))
            changed.append('пароль')

        if changed:
            flash(f'Успішно змінено: {", ".join(changed)}.', 'success')
        else:
            flash('Нічого не змінено.', 'info')

        # Повертаємо на відповідний список
        if user['role'] == 'student':
            return redirect(url_for('admin_students'))
        elif user['role'] == 'teacher':
            return redirect(url_for('admin_teachers'))
        else:
            return redirect(url_for('admin_dashboard'))

    return render_template('admin/edit_credentials.html', user=user)



# ============================================================
# КЕРІВНИК ПРАКТИКИ (TEACHER)
# ============================================================

@app.route('/teacher')
@teacher_required
def teacher_dashboard():
    """Головна панель керівника — список його студентів."""
    user_id = get_current_user_id()
    teacher = get_teacher_by_user_id(user_id)

    if not teacher:
        flash('Профіль керівника не знайдено.', 'danger')
        return redirect(url_for('logout'))

    practices = get_practices_by_teacher_id(teacher['id'])

    # Дані для графіка статусів
    chart_data = {
        'labels': ['Призначено', 'В процесі', 'Звіт здано', 'Оцінено'],
        'values': [
            sum(1 for p in practices if p['status'] == 'assigned'),
            sum(1 for p in practices if p['status'] == 'in_progress'),
            sum(1 for p in practices if p['status'] == 'submitted'),
            sum(1 for p in practices if p['status'] == 'graded'),
        ],
    }

    return render_template('teacher/dashboard.html',
                           teacher=teacher,
                           practices=practices,
                           chart_data=chart_data)


@app.route('/teacher/report/<int:practice_id>', methods=['GET', 'POST'])
@teacher_required
def teacher_view_report(practice_id):
    """Перегляд звіту студента та виставлення оцінки."""
    practice = get_practice_by_id(practice_id)
    report   = get_report_by_practice_id(practice_id)
    diary    = get_diary_entries(practice_id)

    if not practice:
        flash('Практику не знайдено.', 'danger')
        return redirect(url_for('teacher_dashboard'))

    # Перевірка: керівник може працювати лише зі СВОЇМИ студентами
    teacher = get_teacher_by_user_id(get_current_user_id())
    if not teacher or practice['teacher_id'] != teacher['id']:
        flash('Немає доступу — це не ваш студент.', 'danger')
        return redirect(url_for('teacher_dashboard'))

    if request.method == 'POST':
        action  = request.form.get('action', 'grade')
        comment = request.form.get('comment', '').strip()

        if action == 'revision':
            # Повернення на доопрацювання — потрібен коментар
            if not comment:
                flash('Вкажіть коментар — що саме потрібно доопрацювати.', 'warning')
            else:
                return_report_for_revision(practice_id, comment)
                flash('Звіт повернено студенту на доопрацювання.', 'info')
                return redirect(url_for('teacher_dashboard'))
        else:
            # Виставлення оцінки
            grade = request.form.get('grade')
            grade_error = v.validate_grade(grade)
            if grade_error:
                flash(grade_error, 'warning')
            else:
                grade_report(practice_id, int(grade), comment)
                flash('Оцінку виставлено!', 'success')
                return redirect(url_for('teacher_dashboard'))

    return render_template('teacher/report.html',
                           practice=practice,
                           report=report,
                           diary=diary)


# ============================================================
# СТУДЕНТ
# ============================================================

@app.route('/student')
@student_required
def student_dashboard():
    """Головна панель студента — інформація про практику."""
    user_id = get_current_user_id()
    student  = get_student_by_user_id(user_id)

    if not student:
        flash('Профіль студента не знайдено.', 'danger')
        return redirect(url_for('logout'))

    practice = get_practice_by_student_id(student['id'])
    report   = get_report_by_practice_id(practice['id']) if practice else None

    return render_template('student/dashboard.html',
                           student=student,
                           practice=practice,
                           report=report)


@app.route('/student/diary', methods=['GET', 'POST'])
@student_required
def student_diary():
    """Щоденник практики — перегляд і додавання записів."""
    user_id  = get_current_user_id()
    student  = get_student_by_user_id(user_id)
    practice = get_practice_by_student_id(student['id'])

    if not practice:
        flash('Вам ще не призначено практику.', 'warning')
        return redirect(url_for('student_dashboard'))

    if request.method == 'POST':
        entry_date = request.form.get('entry_date')
        content    = request.form.get('content', '').strip()

        content_error = v.validate_text(content, "Зміст запису", min_len=3, max_len=2000)
        if not entry_date:
            flash('Вкажіть дату запису.', 'warning')
        elif content_error:
            flash(content_error, 'warning')
        else:
            add_diary_entry(practice['id'], entry_date, content)
            # ✅ АВТОСТАТУС — якщо практика ще 'assigned', змінюємо на 'in_progress'
            if practice['status'] == 'assigned':
                update_practice_status(practice['id'], 'in_progress')
            flash('Запис додано!', 'success')
            return redirect(url_for('student_diary'))

    entries = get_diary_entries(practice['id'])
    return render_template('student/diary.html',
                           practice=practice,
                           entries=entries)


@app.route('/student/diary/delete/<int:entry_id>')
@student_required
def student_delete_diary_entry(entry_id):
    """Видалення запису щоденника."""
    delete_diary_entry(entry_id)
    flash('Запис видалено.', 'info')
    return redirect(url_for('student_diary'))


@app.route('/student/report', methods=['GET', 'POST'])
@student_required
def student_report():
    """Здача фінального звіту."""
    user_id  = get_current_user_id()
    student  = get_student_by_user_id(user_id)
    practice = get_practice_by_student_id(student['id'])

    if not practice:
        flash('Вам ще не призначено практику.', 'warning')
        return redirect(url_for('student_dashboard'))

    if request.method == 'POST':
        content_text = request.form.get('content_text', '').strip()

        # Обробка завантаженого файлу (необов'язкового)
        file_path = None
        existing = get_report_by_practice_id(practice['id'])
        if existing and existing['file_path']:
            file_path = existing['file_path']  # зберігаємо старий файл якщо новий не завантажили

        uploaded = request.files.get('report_file')
        if uploaded and uploaded.filename:
            if not allowed_report_file(uploaded.filename):
                flash('Дозволені формати файлу: PDF, Word, зображення, ZIP.', 'danger')
                report = get_report_by_practice_id(practice['id'])
                return render_template('student/report.html', practice=practice, report=report)
            # Зберігаємо з унікальною назвою: report_<practice_id>_<оригінальна назва>
            safe_name = secure_filename(uploaded.filename)
            file_path = f"report_{practice['id']}_{safe_name}"
            uploaded.save(os.path.join(app.config['REPORT_FOLDER'], file_path))

        # Має бути хоча б текст АБО файл
        if not content_text and not file_path:
            flash('Додайте текст звіту або прикріпіть файл.', 'warning')
        else:
            submit_report(practice['id'], content_text, file_path)
            flash('Звіт успішно здано! Очікуйте перевірки.', 'success')
            return redirect(url_for('student_dashboard'))

    report = get_report_by_practice_id(practice['id'])
    return render_template('student/report.html',
                           practice=practice,
                           report=report)


@app.route('/report/download/<int:practice_id>')
@login_required
def download_report_file(practice_id):
    """Завантаження прикріпленого файлу звіту (студент свій, керівник/адмін будь-який)."""
    report = get_report_by_practice_id(practice_id)
    if not report or not report['file_path']:
        flash('Файл не знайдено.', 'danger')
        return redirect(url_for('dashboard'))

    # Перевірка доступу: студент може качати тільки свій звіт
    role = get_current_user_role()
    if role == 'student':
        student = get_student_by_user_id(get_current_user_id())
        practice = get_practice_by_id(practice_id)
        if not student or not practice or practice['student_id'] != student['id']:
            flash('Немає доступу до цього файлу.', 'danger')
            return redirect(url_for('student_dashboard'))

    filepath = os.path.join(app.config['REPORT_FOLDER'], report['file_path'])
    if not os.path.exists(filepath):
        flash('Файл відсутній на сервері.', 'danger')
        return redirect(url_for('dashboard'))

    return send_file(filepath, as_attachment=True)




# ============================================================
# МІЙ АКАУНТ — для всіх ролей
# ============================================================

@app.route('/account', methods=['GET', 'POST'])
@login_required
def my_account():
    """Зміна логіну та пароля для будь-якої ролі."""
    user_id = get_current_user_id()
    from database import get_user_by_id
    user = get_user_by_id(user_id)

    if request.method == 'POST':
        new_login    = request.form.get('login', '').strip()
        old_password = request.form.get('old_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        changed = []

        # Перевіряємо старий пароль якщо хочуть змінити пароль
        if new_password and not old_password:
            flash('Щоб змінити пароль — введіть спочатку старий.', 'warning')
            return render_template('account.html', user=user)

        if old_password and not check_password(old_password, user['password_hash']):
            flash('Старий пароль введено невірно.', 'danger')
            return render_template('account.html', user=user)

        # Змінюємо логін
        if new_login and new_login != user['login']:
            err = v.validate_login(new_login)
            if err:
                flash(err, 'warning')
                return render_template('account.html', user=user)
            ok = update_user_login(user_id, new_login)
            if ok:
                session['user_login'] = new_login
                changed.append('логін')
            else:
                flash(f'Логін «{new_login}» вже зайнятий.', 'danger')
                return render_template('account.html', user=user)

        # Змінюємо пароль
        if new_password:
            err = v.validate_password(new_password)
            if err:
                flash(err, 'warning')
                return render_template('account.html', user=user)
            update_user_password(user_id, hash_password(new_password))
            changed.append('пароль')

        if changed:
            flash(f'Успішно змінено: {", ".join(changed)}!', 'success')
        else:
            flash('Нічого не змінено.', 'info')

        return redirect(url_for('my_account'))

    return render_template('account.html', user=user)


@app.route('/account/avatar', methods=['POST'])
@login_required
def upload_avatar():
    """Завантаження аватарки користувача."""
    user_id = get_current_user_id()

    if 'avatar' not in request.files:
        flash('Файл не обрано.', 'warning')
        return redirect(url_for('my_account'))

    file = request.files['avatar']

    if file.filename == '':
        flash('Файл не обрано.', 'warning')
        return redirect(url_for('my_account'))

    if not allowed_file(file.filename):
        flash('Дозволені формати: PNG, JPG, GIF, WEBP.', 'danger')
        return redirect(url_for('my_account'))

    # Зберігаємо файл з унікальною назвою (user_id + розширення)
    ext = file.filename.rsplit('.', 1)[1].lower()
    filename = f'avatar_{user_id}.{ext}'
    filepath = os.path.join(app.config['AVATAR_FOLDER'], filename)
    file.save(filepath)

    # Зберігаємо назву файлу в БД
    update_user_avatar(user_id, filename)
    session['user_avatar'] = filename  # Оновлюємо сесію щоб аватарка одразу з'явилась у сайдбарі
    flash('Аватарку успішно оновлено! 🎉', 'success')
    return redirect(url_for('my_account'))


@app.route('/account/avatar/delete', methods=['POST'])
@login_required
def delete_avatar():
    """Видалення аватарки — повертає першу літеру логіну."""
    user_id = get_current_user_id()
    from database import get_user_by_id
    user = get_user_by_id(user_id)

    # Видаляємо файл з диску якщо він є
    if user and user['avatar']:
        old_path = os.path.join(app.config['AVATAR_FOLDER'], user['avatar'])
        if os.path.exists(old_path):
            try:
                os.remove(old_path)
            except OSError:
                pass  # якщо не вдалось видалити файл — не критично

    # Очищаємо аватарку в БД та сесії
    update_user_avatar(user_id, None)
    session['user_avatar'] = ''
    flash('Аватарку видалено.', 'info')
    return redirect(url_for('my_account'))


@app.route('/account/avatar/privacy', methods=['POST'])
@login_required
def toggle_avatar_privacy():
    """Перемикає приватність аватарки."""
    from database import update_avatar_privacy
    user_id = get_current_user_id()
    # checkbox: якщо приходить 'on' — приватна, інакше — публічна
    is_private = request.form.get('avatar_private') == 'on'
    update_avatar_privacy(user_id, is_private)
    if is_private:
        flash('Аватарку приховано від інших користувачів.', 'info')
    else:
        flash('Аватарку тепер видно іншим користувачам.', 'success')
    return redirect(url_for('my_account'))



# ============================================================
# ЕКСПОРТ ДАНИХ — EXCEL та PDF
# ============================================================

def _build_excel(practices, title_text):
    """Створює Excel-файл зі списку практик. Повертає BytesIO."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Практики"

    headers = ['№', 'Студент', 'Група', 'База практики', 'Керівник',
               'Тема', 'Початок', 'Кінець', 'Статус']
    ws.append(headers)

    header_fill = PatternFill(start_color='4F8EF7', end_color='4F8EF7', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    status_map = {
        'assigned': 'Призначено', 'in_progress': 'В процесі',
        'submitted': 'Звіт здано', 'graded': 'Оцінено'
    }

    for i, p in enumerate(practices, start=1):
        ws.append([
            i, p['student_name'], p['student_group'], p['company_name'],
            p['teacher_name'], p['topic'], p['start_date'], p['end_date'],
            status_map.get(p['status'], p['status']),
        ])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='center')

    widths = [5, 25, 12, 28, 22, 35, 12, 12, 14]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _build_pdf(practices, title_text):
    """Створює PDF-файл зі списку практик. Повертає BytesIO."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = 'Helvetica'
    try:
        for font_path in [
            'C:/Windows/Fonts/arial.ttf',
            'C:/Windows/Fonts/calibri.ttf',
            '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
        ]:
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('CustomFont', font_path))
                font_name = 'CustomFont'
                break
    except Exception:
        font_name = 'Helvetica'

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1*cm, rightMargin=1*cm)

    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Title'],
                                  fontName=font_name, fontSize=16)
    date_style = ParagraphStyle('Date', parent=styles['Normal'],
                                 fontName=font_name, fontSize=9,
                                 textColor=colors.grey)

    elements.append(Paragraph(title_text, title_style))
    elements.append(Paragraph(
        f'Дата формування: {datetime.now().strftime("%d.%m.%Y %H:%M")}', date_style))
    elements.append(Spacer(1, 0.5*cm))

    status_map = {
        'assigned': 'Призначено', 'in_progress': 'В процесі',
        'submitted': 'Звіт здано', 'graded': 'Оцінено'
    }

    cell_style = ParagraphStyle('Cell', fontName=font_name, fontSize=8, leading=10)
    head_style = ParagraphStyle('Head', fontName=font_name, fontSize=9,
                                 textColor=colors.white, leading=11)

    def P(text, style=cell_style):
        return Paragraph(str(text), style)

    data = [[
        P('№', head_style), P('Студент', head_style), P('Група', head_style),
        P('База практики', head_style), P('Керівник', head_style),
        P('Тема', head_style), P('Статус', head_style)
    ]]

    for i, p in enumerate(practices, start=1):
        data.append([
            P(i), P(p['student_name']), P(p['student_group']),
            P(p['company_name']), P(p['teacher_name']),
            P(p['topic']), P(status_map.get(p['status'], p['status']))
        ])

    table = Table(data, colWidths=[1*cm, 4.5*cm, 2.2*cm, 5*cm, 4*cm, 6.5*cm, 2.5*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F8EF7')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CCCCCC')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F7FA')]),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)

    doc.build(elements)
    output.seek(0)
    return output


# --- Експорт для адміна (усі практики) ---

@app.route('/admin/export/excel')
@admin_required
def export_excel():
    """Адмін експортує усі практики в Excel."""
    practices = get_all_practices()
    output = _build_excel(practices, 'EduSync — Звіт по практиках')
    filename = f'practices_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)


@app.route('/admin/export/pdf')
@admin_required
def export_pdf():
    """Адмін експортує усі практики в PDF."""
    practices = get_all_practices()
    output = _build_pdf(practices, 'EduSync — Звіт по практиках')
    filename = f'practices_{datetime.now().strftime("%Y-%m-%d")}.pdf'
    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


# --- Експорт для викладача (тільки свої студенти) ---

@app.route('/teacher/export/excel')
@teacher_required
def teacher_export_excel():
    """Викладач експортує практики своїх студентів в Excel."""
    teacher = get_teacher_by_user_id(get_current_user_id())
    if not teacher:
        flash('Профіль керівника не знайдено.', 'danger')
        return redirect(url_for('teacher_dashboard'))
    practices = get_practices_by_teacher_id(teacher['id'])
    output = _build_excel(practices, 'EduSync — Мої студенти')
    filename = f'my_students_{datetime.now().strftime("%Y-%m-%d")}.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename)


@app.route('/teacher/export/pdf')
@teacher_required
def teacher_export_pdf():
    """Викладач експортує практики своїх студентів в PDF."""
    teacher = get_teacher_by_user_id(get_current_user_id())
    if not teacher:
        flash('Профіль керівника не знайдено.', 'danger')
        return redirect(url_for('teacher_dashboard'))
    practices = get_practices_by_teacher_id(teacher['id'])
    output = _build_pdf(practices, 'EduSync — Мої студенти')
    filename = f'my_students_{datetime.now().strftime("%Y-%m-%d")}.pdf'
    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


# ============================================================
# ЗАПУСК СЕРВЕРА
# ============================================================

if __name__ == '__main__':
    # Створюємо таблиці при першому запуску
    init_db()
    add_avatar_column()  # Додаємо колонку аватарки якщо ще немає
    from database import migrate_status_check
    migrate_status_check()  # Дозволяємо статус 'revision' для наявних баз

    # Створюємо адміністратора за замовчуванням (якщо потрібно)
    from database import get_user_by_login, create_user
    if get_user_by_login('admin') is None:
        create_user('admin', hash_password('admin123'), 'admin')
        print("✅ Створено адміністратора: login=admin, password=admin123")
        print("⚠️  Змініть пароль після першого входу!")

    print("🚀 Сервер запущено: http://localhost:5000")

    # debug=True — автоматично перезапускає сервер при змінах у коді
    app.run(debug=True)